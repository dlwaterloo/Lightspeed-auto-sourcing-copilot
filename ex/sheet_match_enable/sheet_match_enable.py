import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment


def read_leftmost_column_to_list(file_path, sheet_name=0):
	"""
	从指定的Excel文件中读取最左侧一列的正则表达式模式到列表中。

	参数:
	- file_path: Excel文件的路径。
	- sheet_name: 要读取的工作表名称或索引，默认为第一个工作表。

	返回:
	- regex_patterns: 包含所有正则表达式的列表。
	"""
	# 读取Excel文件，只读取最左侧的一列
	df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=[0], header=None)

	# 提取最左侧列的数据，并过滤掉任何可能存在的NaN值
	regex_patterns = df.iloc[:, 0].dropna().astype(str).tolist()

	return regex_patterns


def parse_string_to_dict(input_string):
	# 定义分割符和输出字典
	delimiters = ' - | – '
	result_dict = {}
	try:
		# 分割输入字符串为各个部分
		parts = re.split(delimiters, input_string)

		# 解析日期
		date_part = parts[0].strip()
		parsed_date = pd.to_datetime(date_part, errors='raise')
		result_dict['date'] = parsed_date

		# 解析轮次
		round_part = parts[1].strip()
		result_dict['round'] = round_part

		# 解析金额
		amount_part = parts[2].strip()
		result_dict['amount'] = amount_part

		if len(parts) >= 4:
			# 解析来源公司列表
			from_part = parts[3].strip().replace('from(', '').replace(')', '')
			companies_list = [company.strip().replace('领投', '').replace('跟投', '') for company in from_part.split('/')]
			result_dict['from_companies'] = companies_list
		else:
			result_dict['from_companies'] = []
		return result_dict

	except Exception as e:
		print(f"An error occurred while parsing {input_string}: {e}")
		return None


def check_peer_match(test_string, peers):
	for peer in peers:
		if re.match(peer, test_string):
			return True
	return False


def check_conditions(info, result, conditions):
	fund_st, hist_st, peer_st = info
	fund_res, hist_res, peer_res = result
	ans = False
	for c in conditions:
		fund_enable, hist_enable, peer_enable = c
		ans = ans or (
			(fund_res >= fund_st if fund_enable else True) and
			(hist_res >= hist_st if hist_enable else True) and
			(peer_res >= peer_st if peer_enable else True)
		)
	return ans


def check_and_collect(row, column_name, info, peers, conditions):
	year, all_funds_count, funding_hist_count, peer_funds_count = info
	year = str(year)
	start_date = year + '-01-01'
	end_date = year + '-12-31'
	hist_count = 0
	funds_count = 0
	contain_peer_count = 0
	try:
		hists = row[column_name].split('\n')
		for h in hists:
			hist_dict = parse_string_to_dict(h)
			if hist_dict and hist_dict['date'] >= pd.to_datetime(start_date) and hist_dict['date'] <= pd.to_datetime(end_date):
				from_list = hist_dict['from_companies']
				hist_count = hist_count + 1
				funds_count = funds_count + len(from_list)
				contain_peer_count = sum(check_peer_match(str, peers) for str in from_list)
				# print(hist_dict)
	except Exception as e:
		print(f"An error occurred while parsing {row[column_name]}: {e}")
		return False

	info = all_funds_count, funding_hist_count, peer_funds_count
	result = funds_count, hist_count, contain_peer_count
	return check_conditions(info, result, conditions)

def filter_timestamp(df, info, peers, conditions):

	column = 'Funding History'

	df_filtered = df[df.apply(check_and_collect, args=(
		column, info, peers, conditions), axis=1)].reset_index(drop=True)
	# 显示数据概览
	print(df_filtered)

	return df_filtered


# 输入追踪机构表
tracked_file_path = 'PF Tracked List For Match.xlsx'  # 替换为你的文件路径
regex_patterns = read_leftmost_column_to_list(tracked_file_path)
# print(regex_patterns)

# 读取配置
cfg_file_path = 'Param_enable.xlsx'
df_cfg = pd.read_excel(cfg_file_path, header=0)

# 获取列名（标题）
column_titles = df_cfg.columns.tolist()
first_data_row = df_cfg.iloc[0]
cfg_tuple = tuple(first_data_row[title] for title in column_titles)
print(cfg_tuple)

# 从第三行开始读取（pandas中index是从0开始，所以第3行是index 2）
# 第二列到第四列对应的是index 1, 2, 3
start_row = 2  # 因为索引从0开始，所以第3行是索引2
col_labels = df_cfg.columns[[1, 2, 3]]

# 初始化一个空列表来保存三元组
conditions = []

# 遍历DataFrame中的每一行，直到第二列为空
for index, row in df_cfg.iloc[start_row:].iterrows():
    if pd.isna(row[col_labels[0]]):  # 如果第二列的值为空，则停止循环
        break
    # 创建一个三元组，并添加到列表中
    triplet = tuple(row[col] for col in col_labels)
    conditions.append(triplet)

print(conditions)

# 读取Excel文件
input_file = 'Peer Funds 20250120-New Investment.xlsx'  # 输入文件名
output_file = 'Filtered Peer Funds 20250120-New Investment.xlsx'  # 输出文件名

df = pd.read_excel(input_file)

df_hist = filter_timestamp(df, cfg_tuple, regex_patterns, conditions)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
	df_hist.to_excel(writer, sheet_name='New Investments', index=True)

	# 获取工作簿和工作表
	workbook = writer.book
	worksheet = writer.sheets['New Investments']

	# 调整行高以适应内容长度
	for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
		max_height = 0
		for cell in row:
			# 计算单元格所需的最大高度
			lines = str(cell.value).split('\n')
			max_lines = max([len(line) // 60 + 1 for line in lines])  # 假设每行最多显示60个字符
			cell_height = max_lines * 15  # 每行大约15个单位的高度
			if cell_height > max_height:
				max_height = cell_height

		# 设置自动换行
		# 创建一次 Alignment 对象
		# align = Alignment(wrap_text=True)

		# 使用同一个 Alignment 对象设置所有单元格
		# for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
		#	for cell in row:
		#		cell.alignment = align

		# 调整列宽以适应内容长度
		for column_cells in worksheet.columns:
			length = max(max([min(len(line) + 1, 61) for line in str(cell.value).split('\n')]) for cell in column_cells if cell.value is not None)
			worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2  # 添加一些额外的空间

		# 调整行高以适应内容长度
		for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
			max_height = 0
			for cell in row:
				if isinstance(cell.value, str):
					lines = cell.value.split('\n')
					max_lines = max([len(line) // 60 + 1 for line in lines])  # 假设每行最多显示60个字符
					cell_height = max_lines * 15  # 每行大约15个单位的高度
					if cell_height > max_height:
						max_height = cell_height

			# 设置行高
			worksheet.row_dimensions[row[0].row].height = max_height

print(f"格式化完成，结果已保存到 {output_file}")

# df_hist.to_excel(output_file, index=True)
